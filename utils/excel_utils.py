import os
from openpyxl import Workbook, load_workbook


def create_excel_file():
    """
    Creates an Excel file with a predefined structure inside the 'data' folder.

    Args:
        None

    Returns:
        None

    Raises:
        IOError: If there is an error saving the file.
    """
    # Get the absolute path to the 'data' folder
    current_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(current_dir, "../data")
    os.makedirs(data_dir, exist_ok=True)

    # Define the path for the Excel file
    file_path = os.path.join(data_dir, "udemy_courses.xlsx")

    print("filepath", file_path)

    # Create a new workbook and set the active worksheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Udemy Courses"

    # Add headers to the sheet
    sheet.append(
        [
            "Category",
            "Subcategory",
            "Subcategory URL",
            "Course Title",
            "Course URL",
            "Description",
            "Rating",
            "Number of Reviews",
            "Total Hours",
            "Number of Lectures",
            "Level",
            "Current Price",
            "Original Price",
            "Discount",
        ]
    )

    # Save the workbook to the 'data' folder
    try:
        workbook.save(file_path)
        print(f"Excel file created at: {file_path}")
    except IOError as e:
        print(f"Error saving Excel file: {e}")


def write_courses_to_excel(category_name, sub_category_name, subcategory_url, courses):
    """
    Writes course data to an Excel file inside the 'data' folder.

    Args:
        category_name (str): The name of the course category.
        sub_category_name (str): The name of the course subcategory.
        subcategory_url (str): The URL of the course subcategory.
        courses (list of dict): A list of dictionaries representing courses, with:
            - 'title' (str): Course title.
            - 'url' (str): Course URL.
            - 'description' (str): Course description.
            - 'rating' (float): Course rating.
            - 'num_reviews' (int): Number of reviews.
            - 'total_hours' (float): Total course hours.
            - 'num_lectures' (int): Number of lectures.
            - 'level' (str): Course difficulty level.

    Returns:
        None
    """
    if not courses:
        print(f"No courses found for subcategory: {sub_category_name}, skipping...")
        return

    current_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(current_dir, "../data")
    os.makedirs(data_dir, exist_ok=True)

    file_path = os.path.join(data_dir, "udemy_courses.xlsx")

    if not os.path.exists(file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Courses"

        sheet.append(
            [
                "Category Name",
                "Subcategory Name",
                "Subcategory URL",
                "Course Title",
                "Course URL",
                "Description",
                "Rating",
                "Number of Reviews",
                "Total Hours",
                "Number of Lectures",
                "Level",
                "Current Price",
                "Original Price",
                "Discount",
            ]
        )

        workbook.save(file_path)
        print(f"Created new Excel file at: {file_path}")

    workbook = load_workbook(file_path)
    sheet = workbook.active

    for course in courses:
        sheet.append(
            [
                category_name,
                sub_category_name,
                subcategory_url,
                course.get("title", "N/A"),
                course.get("url", "N/A"),
                course.get("description", "N/A"),
                course.get("rating", "N/A"),
                course.get("num_reviews", "N/A"),
                course.get("total_hours", "N/A"),
                course.get("num_lectures", "N/A"),
                course.get("level", "N/A"),
                course.get("current_price", "N/A"),
                course.get("original_price", "N/A"),
                course.get("discount", "N/A"),
            ]
        )

    try:
        workbook.save(file_path)
        print(f"Data for {sub_category_name} saved to {file_path}")
    except IOError as e:
        print(f"Error saving Excel file: {e}")
